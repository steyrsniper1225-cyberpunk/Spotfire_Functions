import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import base64
from PIL import Image
import io
import os
from tqdm import tqdm

# Linux 클라우드 환경 경로 설정
BASE_DIR = "/data_home/user/2025/username/Python"
INPUT_FILE = os.path.join(BASE_DIR, "Encoded_Data.xlsx")   # Base64가 포함된 원본 엑셀 파일
OUTPUT_FILE = os.path.join(BASE_DIR, "Decoded_Images.xlsx") # 이미지가 삽입된 결과 엑셀 파일

# Base64 문자열이 들어있는 타겟 컬럼명 리스트
TARGET_COLS = [
    '검사_Image_Base64',
    'FIB_Image_Base64',
    'EDS_Image_Base64',
    'Ion_mapping_Base64'
]

def decode_base64_to_openpyxl_image(base64_str):
    """Base64 문자열을 디코딩하여 openpyxl 이미지 객체로 반환"""
    if not base64_str or not isinstance(base64_str, str):
        return None
        
    if "base64," in base64_str:
        base64_str = base64_str.split("base64,")[1]
        
    try:
        img_data = base64.b64decode(base64_str)
        img_buffer = io.BytesIO(img_data)
        
        # PIL을 통해 이미지 모드 확인 및 안전하게 JPEG 버퍼로 재저장
        with Image.open(img_buffer) as pil_img:
            if pil_img.mode != 'RGB':
                pil_img = pil_img.convert('RGB')
            
            out_buffer = io.BytesIO()
            pil_img.save(out_buffer, format="JPEG")
            out_buffer.seek(0) # 버퍼의 시작 위치로 이동
            
            # openpyxl에서 엑셀에 삽입할 수 있는 이미지 객체로 변환
            xl_img = OpenpyxlImage(out_buffer)
            return xl_img
    except Exception as e:
        return None

# 1. 엑셀 워크북 로드
print(f"엑셀 파일 로드 중: {INPUT_FILE}")
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

# 2. 헤더(1행)에서 타겟 컬럼의 엑셀 열 알파벳(예: 'A', 'B') 매핑
col_mapping = {}
for cell in ws[1]:
    if cell.value in TARGET_COLS:
        col_mapping[cell.value] = cell.column_letter

# 3. 데이터 행 순회 (2행부터 마지막 행까지)
max_row = ws.max_row
total_inserted = 0

for row_idx in tqdm(range(2, max_row + 1), desc="이미지 삽입 진행도"):
    for col_name, col_letter in col_mapping.items():
        cell_coord = f"{col_letter}{row_idx}"
        cell = ws[cell_coord]
        
        base64_str = cell.value
        
        if base64_str and isinstance(base64_str, str):
            xl_img = decode_base64_to_openpyxl_image(base64_str)
            
            if xl_img:
                # 엑셀 셀 크기에 맞게 이미지 사이즈 조정이 필요한 경우 아래 주석 해제 후 설정
                # xl_img.width = 150
                # xl_img.height = 150
                
                # 기존의 긴 Base64 텍스트 값은 삭제 (셀을 비움)
                cell.value = None
                
                # 해당 셀 좌표에 이미지 추가
                ws.add_image(xl_img, cell_coord)
                total_inserted += 1

# 4. 결과 저장
print(f"저장 중: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)
wb.close()

print("-" * 50)
print(f"작업 완료. 총 {total_inserted}개의 이미지가 엑셀에 삽입되었습니다.")