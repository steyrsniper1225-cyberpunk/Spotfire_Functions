import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import base64
from PIL import Image
import io
import os
import glob
from tqdm import tqdm

# Linux 클라우드 환경 경로 설정
BASE_DIR = "/data_home/user/2025/username/Python"
OUTPUT_FILE = os.path.join(BASE_DIR, "Spotfire_Ready_All.csv")

# 🎯 추출할 엑셀 이미지 컬럼 지정 (엑셀 열 알파벳 : CSV에 저장될 컬럼명)
TARGET_IMAGE_COLS = {
    'A': '검사_Image_Base64',
    'B': 'FIB_Image_Base64',
    'C': 'EDS_Image_Base64',
    'D': 'Ion_mapping_Base64'
}

def convert_image_to_base64(img_data, target_size=(800, 800)):
    """openpyxl에서 추출한 이미지 바이트 데이터를 Base64 문자열로 변환"""
    with Image.open(io.BytesIO(img_data)) as img:
        if img.mode != 'RGB':
            img = img.convert('RGB')
        img.thumbnail(target_size)
        
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=75)
        img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        return f"data:image/jpeg;base64,{img_str}"

# BASE_DIR 내의 모든 xlsx 파일 검색 (열려있는 임시 파일 제외)
xlsx_files = [f for f in glob.glob(os.path.join(BASE_DIR, "*.xlsx")) if not os.path.basename(f).startswith('~$')]

if os.path.exists(OUTPUT_FILE):
    os.remove(OUTPUT_FILE)

total_rows = 0

for i, file_path in enumerate(xlsx_files):
    file_name = os.path.basename(file_path)
    print(f"처리 중 [{i+1}/{len(xlsx_files)}]: {file_name}")
    
    # 1. 메타데이터 로드 (A~D열 등 이미지 열을 제외하고 읽을 경우 usecols 조정 필요)
    # 현재는 전체 열을 다 읽어오되, 추후 불필요한 열은 drop 할 수 있습니다.
    df = pd.read_excel(file_path)
    
    # 2. Excel 이미지 추출
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # 컬럼별로 이미지 Base64를 저장할 딕셔너리 초기화
    # 예: {'A': {0: 'base...', 1: 'base...'}, 'B': {...}}
    image_mappings = {col: {} for col in TARGET_IMAGE_COLS.keys()}
    
    for image in tqdm(ws._images, desc=f"변환 진행도", unit="장"):
        try:
            # 이미지의 좌측 상단 셀 기준 행/열 인덱스 (0부터 시작)
            col_idx = image.anchor._from.col
            row_idx = image.anchor._from.row
            
            # 0-indexed 열 번호를 엑셀 알파벳(A, B, C...)으로 변환
            col_letter = get_column_letter(col_idx + 1)
            
            # 지정된 타겟 컬럼에 있는 이미지만 처리
            if col_letter in TARGET_IMAGE_COLS:
                img_data = image.ref.getvalue()
                base64_str = convert_image_to_base64(img_data)
                
                # 데이터프레임 인덱스 (엑셀 1행이 헤더인 경우 -1)
                df_idx = row_idx - 1 
                
                # 해당 열, 해당 행에 Base64 문자열 저장
                image_mappings[col_letter][df_idx] = base64_str
                
        except Exception as e:
            print(f"\n[{file_name}] {col_letter}{row_idx+1} 이미지 처리 오류: {e}")
            
    wb.close()
    
    # 3. DataFrame에 각각의 이미지 매핑
    for col_letter, csv_col_name in TARGET_IMAGE_COLS.items():
        df[csv_col_name] = df.index.map(image_mappings[col_letter])
    
    df['Source_File'] = file_name
    
    # 4. CSV 이어쓰기
    write_mode = 'w' if i == 0 else 'a'
    write_header = True if i == 0 else False
    
    df.to_csv(OUTPUT_FILE, mode=write_mode, header=write_header, index=False, encoding='utf-8-sig')
    
    total_rows += len(df)
    print("-" * 50)

if total_rows > 0:
    print(f"모든 파일 처리 완료. 총 {total_rows}행 저장 위치: {OUTPUT_FILE}")
else:
    print("처리할 .xlsx 파일이 없습니다.")